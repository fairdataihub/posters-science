import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('scholarlyArticle_schema_crosswalk.xlsx')
ws = wb.active

# Define SEO priorities based on implementation status
# Properties used in NuxtSchemaPoster get higher priority
seo_priorities = {
    # Core properties being used (HIGH priority)
    '@context': 5,
    '@type': 5,
    'about': 4,
    'abstract': 5,
    'author': 5,
    'citation': 4,
    'datePublished': 4,
    'funder': 4,
    'headline': 4,
    'identifier': 5,
    'image': 4,
    'inLanguage': 3,
    'keywords': 4,
    'license': 3,
    'mainEntityOfPage': 4,
    'name': 5,
    'url': 4,
    'version': 2,
    
    # Properties that could be added (MEDIUM priority)
    'publisher': 4,
    'publicationYear': 3,
    'isPartOf': 3,
    'interactionStatistic': 3,
    'aggregateRating': 3,
    'description': 5,
    
    # Less critical properties (LOW priority)
    'accessibilityAPI': 1,
    'accessibilityControl': 1,
    'accessibilityFeature': 1,
    'accessibilityHazard': 1,
    'accessibilitySummary': 1,
    'accessMode': 1,
    'accessModeSufficient': 1,
    'accountablePerson': 2,
    'acquireLicensePage': 1,
    'additionalType': 2,
    'alternateName': 2,
    'alternativeHeadline': 2,
    'archivedAt': 1,
    'articleBody': 3,
    'articleSection': 2,
    'assesses': 1,
    'associatedMedia': 2,
    'audience': 2,
    'audio': 2,
    'award': 2,
    'backstory': 2,
    'character': 1,
    'comment': 2,
    'commentCount': 2,
    'conditionsOfAccess': 1,
    'contentLocation': 2,
    'contentRating': 1,
    'contentReferenceTime': 1,
    'contributor': 2,
    'copyrightHolder': 2,
    'copyrightNotice': 2,
    'copyrightYear': 2,
    'correction': 1,
    'countryOfOrigin': 1,
    'creativeWorkStatus': 1,
    'creator': 3,
    'creditText': 1,
    'dateCreated': 2,
    'dateModified': 2,
    'digitalSourceType': 1,
    'disambiguatingDescription': 2,
    'discussionUrl': 1,
    'displayLocation': 1,
    'editEIDR': 1,
    'editor': 2,
    'educationalAlignment': 1,
    'educationalLevel': 1,
    'educationalUse': 1,
    'encoding': 2,
    'encodingFormat': 2,
    'exampleOfWork': 1,
    'expires': 1,
    'funding': 3,
    'genre': 2,
    'hasPart': 2,
    'interactivityType': 1,
    'interpretedAsClaim': 1,
    'isAccessibleForFree': 2,
    'isBasedOn': 2,
    'isFamilyFriendly': 1,
    'learningResourceType': 1,
    'locationCreated': 1,
    'mainEntity': 2,
    'maintainer': 1,
    'material': 1,
    'materialExtent': 1,
    'mentions': 2,
    'offers': 1,
    'owner': 2,
    'pageEnd': 1,
    'pageStart': 1,
    'pagination': 1,
    'pattern': 1,
    'position': 1,
    'potentialAction': 1,
    'producer': 2,
    'provider': 1,
    'publication': 1,
    'publisherImprint': 1,
    'publishingPrinciples': 2,
    'recordedAt': 1,
    'releasedEvent': 1,
    'review': 2,
    'sameAs': 3,
    'schemaVersion': 1,
    'sdDatePublished': 1,
    'sdLicense': 1,
    'sdPublisher': 1,
    'size': 1,
    'sourceOrganization': 1,
    'spatial': 1,
    'spatialCoverage': 1,
    'speakable': 1,
    'sponsor': 1,
    'subjectOf': 1,
    'teaches': 1,
    'temporal': 1,
    'temporalCoverage': 1,
    'text': 1,
    'thumbnail': 2,
    'thumbnailUrl': 2,
    'timeRequired': 1,
    'translationOfWork': 1,
    'translator': 1,
    'typicalAgeRange': 1,
    'usageInfo': 1,
    'video': 2,
    'wordCount': 1,
    'workExample': 1,
    'workTranslation': 1,
}

# Update the SEO Priority column (column E, index 4)
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Light yellow
red_fill = PatternFill(start_color='FFB6C6', end_color='FFB6C6', fill_type='solid')  # Light red

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
    schema_prop = row[0].value  # Column A
    if schema_prop:
        priority = seo_priorities.get(schema_prop, 1)
        row[4].value = priority  # Column E (index 4)
        
        # Color code based on priority
        if priority >= 4:
            row[4].fill = green_fill
        elif priority >= 3:
            row[4].fill = yellow_fill
        else:
            row[4].fill = red_fill

# Save the workbook
wb.save('scholarlyArticle_schema_crosswalk.xlsx')
print("✅ Updated SEO Priority (AI estimated) for all properties")
print("\nSummary:")
print(f"  🟢 High Priority (4-5): Properties implemented or critical for SEO")
print(f"  🟡 Medium Priority (3): Useful properties not yet implemented")
print(f"  🔴 Low Priority (1-2): Optional properties")

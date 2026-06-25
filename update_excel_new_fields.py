import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('scholarlyArticle_schema_crosswalk.xlsx')
ws = wb.active

# New properties to add
new_implementations = {
    'publisher': ('poster.value?.publisher', 'Implemented'),
    'interactionStatistic': ('poster.value.likes, poster.value.views', 'Implemented'),
}

# Light green fill for implemented properties
implemented_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

# Update the rows with the new implementations
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
    schema_prop = row[0].value  # Column A
    if schema_prop and schema_prop in new_implementations:
        internal_field, status = new_implementations[schema_prop]
        
        # Update Internal Field (Column B, index 1)
        row[1].value = internal_field
        row[1].fill = implemented_fill
        
        # Update Status (Column C, index 2)
        row[2].value = status
        row[2].fill = implemented_fill

# Save the workbook
wb.save('scholarlyArticle_schema_crosswalk.xlsx')

print("✅ Updated scholarlyArticle_schema_crosswalk.xlsx")
print("\n📊 New implementations added:")
for prop, (field, status) in sorted(new_implementations.items()):
    print(f"  • {prop:30s} → {field:45s} [{status}]")


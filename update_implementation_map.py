import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('scholarlyArticle_schema_crosswalk.xlsx')
ws = wb.active

# Map of schema properties to internal fields and status
implementation_map = {
    '@context': ('N/A', 'Implemented'),
    '@id': ('poster.value?.doi', 'Implemented'),
    '@type': ('N/A', 'Implemented'),
    'about': ('poster.value?.domain', 'Implemented'),
    'abstract': ('poster.value?.description', 'Implemented'),
    'author': ('poster.value?.authors[]', 'Implemented'),
    'citation': ('poster.value.references[]', 'Implemented'),
    'datePublished': ('poster.value?.publishedAt', 'Implemented'),
    'funder': ('poster.value.funding[]', 'Implemented'),
    'headline': ('poster.value?.title', 'Implemented'),
    'identifier': ('poster.value?.doi', 'Implemented'),
    'image': ('poster.value?.imageUrl', 'Implemented'),
    'inLanguage': ('poster.value?.language', 'Implemented'),
    'keywords': ('poster.value?.keywords[]', 'Implemented'),
    'license': ('poster.value?.license', 'Implemented'),
    'mainEntityOfPage': ('poster.value.id', 'Implemented'),
    'name': ('poster.value?.title', 'Implemented'),
    'url': ('poster.value.id', 'Implemented'),
    'version': ('poster.value?.version', 'Implemented'),
}

# Update columns B (Internal Field) and C (Status)
implemented_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # Light green

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
    schema_prop = row[0].value  # Column A
    if schema_prop and schema_prop in implementation_map:
        internal_field, status = implementation_map[schema_prop]
        
        # Update Internal Field (Column B, index 1)
        row[1].value = internal_field
        row[1].fill = implemented_fill
        
        # Update Status (Column C, index 2)
        row[2].value = status
        row[2].fill = implemented_fill

# Save the workbook
wb.save('scholarlyArticle_schema_crosswalk.xlsx')

print("✅ Updated Internal Field and Status columns for implemented properties")
print(f"\n📊 Summary of implemented properties:")
for prop, (field, status) in sorted(implementation_map.items()):
    print(f"  • {prop:25s} → {field:40s} [{status}]")

